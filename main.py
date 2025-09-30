import tkinter as tk
from tkinter import filedialog, messagebox, PhotoImage, font
import pandas as pd
import os
from datetime import datetime
import locale
import platform
import json
import requests, subprocess, time, sys
from packaging import version
import tempfile

# ------------------------------
# Global state 29.09.2025 Janko Jovičić
# ------------------------------
structured_path = None
forklift_number = None  # user-entered value
current_lang = None     # set after root is created

# ------------------------------
# Updater settings
# ------------------------------
UPDATE_URL = "https://codingbyday.github.io/Parts.Master/latest.json"

# ------------------------------
# Translations
# ------------------------------
translations = {
    "en": {
        "upload_label": "Upload structured file to generate the result!",
        "upload_btn": "📂 Upload Structured File",
        "generate_btn": "⚙️ Generate output file",
        "update_btn": "⬇️ Check for Updates",
        "about": "About",
        "footer": "Version {version} | Created by {author}",
        "forklift_prompt": "Enter Forklift name:",
        "file_select_title": "Select structured source file",
        "save_as": "Save As",
        "success": "Success",
        "success_msg": "File saved:\n{path}",
        "error": "Error",
        "error_msg": "Processing failed:\n{err}",
        "no_file": "No file selected",
        "update_available_title": "Update Available",
        "update_available_msg": "A new version {latest} is available!\n\nRelease notes:\n{notes}\n\nDownload and install now?",
        "up_to_date": "You already have the latest version.",
        "update_error": "Could not check for updates:\n{err}",
        "update_failed": "Update Failed",
        "update_ready": "Update",
        "update_ready_msg": "Installer downloaded. It will now run."
    },
    "hr": {
        "upload_label": "Učitaj strukturiranu datoteku za generiranje rezultata!",
        "upload_btn": "📂 Učitaj strukturiranu datoteku",
        "generate_btn": "⚙️ Generiraj izlaznu datoteku",
        "update_btn": "⬇️ Provjeri ažuriranja",
        "about": "O programu",
        "footer": "Verzija {version} | Izradio {author}",
        "forklift_prompt": "Unesite ime viličara:",
        "file_select_title": "Odaberite strukturiranu datoteku",
        "save_as": "Spremi kao",
        "success": "Uspjeh",
        "success_msg": "Datoteka spremljena:\n{path}",
        "error": "Greška",
        "error_msg": "Obrada nije uspjela:\n{err}",
        "no_file": "Nijedna datoteka nije odabrana",
        "update_available_title": "Dostupno ažuriranje",
        "update_available_msg": "Dostupna je nova verzija {latest}!\n\nNapomene uz izdanje:\n{notes}\n\nPreuzeti i instalirati sada?",
        "up_to_date": "Već imate najnoviju verziju.",
        "update_error": "Nije moguće provjeriti ažuriranja:\n{err}",
        "update_failed": "Ažuriranje nije uspjelo",
        "update_ready": "Ažuriranje",
        "update_ready_msg": "Instaler je preuzet. Sada će se pokrenuti."
    }
}

# Try Croatian locale for month names in date/time (fallback to system locale)
try:
    locale.setlocale(locale.LC_TIME, "hr_HR.utf8")
except Exception:
    pass


# ------------------------------
# Utility: bundled resource path
# ------------------------------
def resource_path(relative_path: str) -> str:
    """Get absolute path to resource; works for dev and PyInstaller exe."""
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# ------------------------------
# UI language switching
# ------------------------------
def set_language(lang: str):
    """Switch UI texts between English and Croatian."""
    t = translations[lang]

    title_label.config(text=t["upload_label"])
    upload_btn.config(text=t["upload_btn"])
    generate_btn.config(text=t["generate_btn"])
    update_btn.config(text=t["update_btn"])
    about_btn.config(text=t["about"])
    footer_label.config(text=t["footer"].format(version=APP_VERSION, author=APP_AUTHOR))

    # Reset file label text only if nothing is selected yet
    if structured_path is None:
        file_label.config(text=t["no_file"])

    current_lang.set(lang)
    update_lang_buttons()


def update_lang_buttons():
    # reset both
    btn_en.config(relief="flat", bd=0, bg="#f5f5f5", activebackground="#f5f5f5")
    btn_hr.config(relief="flat", bd=0, bg="#f5f5f5", activebackground="#f5f5f5")
    # highlight selected
    if current_lang.get() == "en":
        btn_en.config(relief="solid", bd=2, bg="#e8f0ff", activebackground="#e8f0ff")
    else:
        btn_hr.config(relief="solid", bd=2, bg="#e8f0ff", activebackground="#e8f0ff")


# ------------------------------
# Updater
# ------------------------------
def check_for_updates():
    t = translations[current_lang.get()]
    try:
        url = f"{UPDATE_URL}?t={int(time.time())}"
        r = requests.get(url, timeout=5, headers={"Cache-Control": "no-cache"})
        r.raise_for_status()
        data = r.json()

        latest = data["version"].strip()
        download_url = data["url"]
        notes = data.get("notes", "")

        if version.parse(latest) > version.parse(APP_VERSION.strip()):
            msg = t["update_available_msg"].format(latest=latest, notes=notes)
            if messagebox.askyesno(t["update_available_title"], msg):
                download_and_install(download_url)
        else:
            messagebox.showinfo(t["update_ready"], t["up_to_date"])
    except Exception as e:
        messagebox.showerror(t["update_failed"], t["update_error"].format(err=e))


def download_and_install(download_url: str):
    t = translations[current_lang.get()]
    try:
        # Save to Windows temp folder
        temp_dir = tempfile.gettempdir()
        local_filename = os.path.join(temp_dir, os.path.basename(download_url))

        with requests.get(download_url, stream=True, timeout=30) as r:
            r.raise_for_status()
            with open(local_filename, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)

        # Notify user
        messagebox.showinfo(t["update_ready"], t["update_ready_msg"])

        # Detect installer type
        if local_filename.lower().endswith(".msi"):
            subprocess.Popen(['msiexec', '/i', local_filename], shell=True)
        else:
            subprocess.Popen([local_filename], shell=True)

        os._exit(0)  # exit app before installer runs
    except Exception as e:
        messagebox.showerror(t["update_failed"], str(e))


# ------------------------------
# Forklift input dialog
# ------------------------------
def ask_forklift_number(parent, app_icon=None):
    t = translations[current_lang.get()]
    dialog = tk.Toplevel(parent)
    dialog.title("Forklift")
    dialog.geometry("400x180")
    dialog.transient(parent)
    dialog.grab_set()

    # set same icon
    if app_icon:
        try:
            dialog.iconbitmap(app_icon)
        except Exception:
            try:
                dialog.iconphoto(False, PhotoImage(file=app_icon))
            except Exception:
                pass

    parent.update_idletasks()
    pw, ph = parent.winfo_width(), parent.winfo_height()
    px, py = parent.winfo_rootx(), parent.winfo_rooty()
    w, h = 400, 180
    x = px + (pw - w) // 2
    y = py + (ph - h) // 2
    dialog.geometry(f"{w}x{h}+{x}+{y}")

    tk.Label(dialog, text=t["forklift_prompt"], font=("Segoe UI", 11)).pack(pady=15)

    entry = tk.Entry(dialog, font=("Segoe UI", 12))
    entry.pack(pady=5)
    entry.focus()

    value = {"result": None}

    def submit():
        value["result"] = entry.get().strip()
        dialog.destroy()

    tk.Button(
        dialog, text="OK", command=submit,
        bg="#2196F3", fg="white", padx=15, pady=5, cursor="hand2"
    ).pack(pady=15)

    parent.wait_window(dialog)
    return value["result"] or "Unknown"


# ------------------------------
# Helpers for BOM logic (from your working version)
# ------------------------------
def item_key(s: str):
    """Natural sort key for 'Item' like '4', '4.3', '4.10.2'."""
    parts = str(s).split(".")
    out = []
    for p in parts:
        p = p.strip()
        if p.isdigit():
            out.append(int(p))
        else:
            digits = "".join(c for c in p if c.isdigit())
            out.append(int(digits) if digits else 0)
    return tuple(out)


def get_direct_children(df: pd.DataFrame, parent_item: str) -> pd.DataFrame:
    """Return only direct children of parent_item (exactly one dot deeper), sorted."""
    p = str(parent_item).strip()
    children = df[df["Item"].apply(
        lambda x: isinstance(x, str) and x.startswith(p + ".") and x.count(".") == p.count(".") + 1
    )].copy()
    if not children.empty:
        children.sort_values(by="Item", key=lambda c: c.map(item_key), inplace=True)
    return children


def normalize_str(x):
    s = "" if pd.isna(x) else str(x)
    return s.strip()


def fmt_revision(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    # convert floats like "1.0" -> "1"
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s


def best_title_from_filename(path: str) -> str:
    """Derive a friendly BOM title from the file name."""
    stem = os.path.splitext(os.path.basename(path))[0]
    stem = stem.replace("structured", "").replace("Structured", "").strip("- _")
    return stem or "BOM"


def write_section(writer, title: str, table: pd.DataFrame, row_pos: int, cols: list) -> int:
    """Write a section title and a table; return new row position."""
    pd.DataFrame([[title]]).to_excel(writer, index=False, header=False, startrow=row_pos)
    row_pos += 2
    if not table.empty:
        table[cols].to_excel(writer, index=False, startrow=row_pos)
        row_pos += len(table) + 3
    else:
        pd.DataFrame([["(no rows)"]]).to_excel(writer, index=False, header=False, startrow=row_pos)
        row_pos += 3
    return row_pos


# ------------------------------
# File upload + prompt forklift
# ------------------------------
def upload_structured():
    global structured_path, forklift_number
    t = translations[current_lang.get()]

    structured_path = filedialog.askopenfilename(
        title=t["file_select_title"],
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if structured_path:
        file_label.config(text=os.path.basename(structured_path))
        forklift_number = ask_forklift_number(root, resource_path("assets/official-logo.ico"))


# ------------------------------
# Main processing (exact logic preserved, with Save As + forklift in 2 places)
# ------------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font, Border
from openpyxl.utils import get_column_letter

# ------------------------------
# Main processing (XlsxWriter with strings_to_numbers)
# ------------------------------
def process_file():
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Border

    t = translations[current_lang.get()]

    if not structured_path:
        messagebox.showerror(
            t["error"],
            t["error_msg"].format(err="Please upload the structured file first.")
        )
        return

    # Save As dialog
    out_path = filedialog.asksaveasfilename(
        title=t["save_as"],
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="10011.xlsx"
    )
    if not out_path:
        return

    # --- Helper for cleaning columns ---
    def clean_column_for_excel(series: pd.Series) -> pd.Series:
        cleaned = []
        for val in series:
            if pd.isna(val):
                cleaned.append("")
                continue
            s = str(val).strip()
            if s.isdigit():   # ✅ pure digits → convert to int
                cleaned.append(int(s))
            else:             # ✅ keep as string (preserve underscores etc.)
                cleaned.append(s)
        return pd.Series(cleaned, index=series.index)

    try:
        # Load and normalize
        df = pd.read_excel(structured_path, dtype={"Item": str})
        df["Item"] = df["Item"].astype(str).str.strip()

        col_mapping = {
            "QTY": "Quantity",
            "Part Number": "Part Number",
            "Component Type": "Type",
            "Filename": "Nomenclature",
            "REV": "Revision",
            "Description": "Product Description",
        }
        df = df.rename(columns=col_mapping)

        req_cols = ["Quantity", "Part Number", "Type", "Nomenclature", "Revision", "Product Description"]
        for c in req_cols:
            if c not in df.columns:
                df[c] = ""

        df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
        df["Type"] = df["Type"].astype(str).str.strip()

        # Normalize
        df["Part Number"] = df["Part Number"].map(normalize_str)
        df["Revision"] = df["Revision"].map(fmt_revision)
        df["Product Description"] = df["Product Description"].map(normalize_str)

        # ✅ Clean key columns for Excel warnings
        df["Part Number"] = clean_column_for_excel(df["Part Number"])
        df["Nomenclature"] = clean_column_for_excel(df["Nomenclature"])
        df["Revision"] = clean_column_for_excel(df["Revision"])

        # Top-level rows (Item with no dot)
        df_parents = df[df["Item"].str.match(r"^\d+$")].copy()
        if not df_parents.empty:
            df_parents.sort_values(by="Item", key=lambda c: c.map(item_key), inplace=True)

        # Write file using xlsxwriter
        with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
            r = 0

            # Date
            now = datetime.now()
            cro_dt = now.strftime("%d. %B %Y. %H:%M:%S")
            pd.DataFrame([[cro_dt]]).to_excel(writer, index=False, header=False, startrow=r); r += 2

            # BOM title
            title_value = forklift_number if forklift_number else best_title_from_filename(structured_path)
            pd.DataFrame([[f"Bill of Material: {title_value}"]]).to_excel(
                writer, index=False, header=False, startrow=r
            ); r += 1

            # Parent table
            if not df_parents.empty:
                df_parents[req_cols].to_excel(writer, index=False, startrow=r)
                r += len(df_parents) + 2
            else:
                pd.DataFrame([["(No top-level items found)"]]).to_excel(
                    writer, index=False, header=False, startrow=r
                ); r += 2

            # ===== Child BOM tables =====
            level = [(str(row["Item"]), row["Part Number"])
                     for _, row in df_parents.iterrows()
                     if str(row["Type"]).lower() == "assembly"]

            while level:
                next_level = []
                for parent_item, pn in level:
                    children = get_direct_children(df, parent_item)
                    pd.DataFrame([[f"Bill of Material: {pn}"]]).to_excel(
                        writer, index=False, header=False, startrow=r
                    ); r += 1
                    if not children.empty:
                        children[req_cols].to_excel(writer, index=False, startrow=r)
                        r += len(children) + 2
                    else:
                        pd.DataFrame([["(no rows)"]]).to_excel(
                            writer, index=False, header=False, startrow=r
                        ); r += 2
                    if not children.empty:
                        asm_children = children[children["Type"].str.lower() == "assembly"]
                        for _, c in asm_children.iterrows():
                            next_level.append((str(c["Item"]), c["Part Number"]))
                next_level.sort(key=lambda t: item_key(t[0]))
                level = next_level

            # ===== Recapitulation =====
            parts_only = df[df["Type"].str.lower() == "part"].copy()
            different_parts = int(parts_only["Part Number"].nunique())
            qty_by_item = df.groupby("Item")["Quantity"].sum().to_dict()

            items_list = df["Item"].dropna().astype(str).tolist()
            is_parent_map = {
                itm: any(other.startswith(itm + ".") for other in items_list if other != itm)
                for itm in items_list
            }
            leaf_mask = df["Item"].map(lambda x: not is_parent_map.get(x, False))
            leaf_parts = df[leaf_mask & (df["Type"].str.lower() == "part")].copy()

            def ancestors(itm: str):
                segs = itm.split(".")
                return [".".join(segs[:i]) for i in range(len(segs)-1, 0, -1)]

            exploded_total = 0.0
            per_pn_qty = {}
            for _, row in leaf_parts.iterrows():
                itm = str(row["Item"])
                pn = row["Part Number"]
                q = float(row["Quantity"])
                mult = 1.0
                for anc in ancestors(itm):
                    mult *= float(qty_by_item.get(anc, 1))
                contrib = q * mult
                exploded_total += contrib
                per_pn_qty[pn] = per_pn_qty.get(pn, 0.0) + contrib

            total_parts = int(round(exploded_total))

            pd.DataFrame([[f"Recapitulation of: {title_value}"]]).to_excel(
                writer, index=False, header=False, startrow=r
            ); r += 1
            pd.DataFrame([
                ["Different parts:", different_parts],
                ["Total parts:", total_parts],
            ]).to_excel(writer, index=False, header=False, startrow=r); r += 3

            # Order & metadata
            first_seen = {}
            seq = 0
            def record_part(pn: str):
                nonlocal seq
                if pn not in first_seen:
                    first_seen[pn] = seq
                    seq += 1
            def dfs_from_item(parent_item: str):
                children = get_direct_children(df, parent_item)
                for _, crow in children.iterrows():
                    if str(crow["Type"]).lower() == "part":
                        record_part(crow["Part Number"])
                    elif str(crow["Type"]).lower() == "assembly":
                        dfs_from_item(str(crow["Item"]))
            for _, prow in df_parents.iterrows():
                if str(prow["Type"]).lower() == "part":
                    record_part(prow["Part Number"])
                elif str(prow["Type"]).lower() == "assembly":
                    dfs_from_item(str(prow["Item"]))

            meta_parts = parts_only[["Part Number", "Revision", "Product Description"]].copy()
            # ✅ Cast to str before calling .str
            meta_parts["_desc_ok"] = meta_parts["Product Description"].astype(str).str.strip().ne("")
            meta_parts["_rev_ok"] = meta_parts["Revision"].astype(str).str.strip().ne("")
            meta_parts.sort_values(by=["_desc_ok", "_rev_ok"], ascending=[False, False], inplace=True)
            meta = meta_parts.groupby("Part Number", as_index=False).first()[["Part Number", "Revision", "Product Description"]]

            rows = []
            for pn, qty in per_pn_qty.items():
                rows.append({
                    "Part Number": pn,
                    "Quantity": int(round(qty)),
                    "_order": first_seen.get(pn, 10**9),
                })
            parts_list = pd.DataFrame(rows)

            if not parts_list.empty:
                parts_list = parts_list.merge(meta, on="Part Number", how="left")
                parts_list.sort_values(by=["_order"], inplace=True)
                parts_list.drop(columns=["_order"], inplace=True)
                parts_list = parts_list[["Quantity", "Part Number", "Revision", "Product Description"]]
                parts_list.to_excel(writer, index=False, startrow=r)
                r += len(parts_list) + 2
            else:
                pd.DataFrame([["(no parts found)"]]).to_excel(writer, index=False, header=False, startrow=r); r += 2

        # === Post-formatting: remove borders + bold on headers ===
        wb = load_workbook(out_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if str(cell.value).strip() in ["Quantity", "Part Number", "Type", "Nomenclature", "Revision", "Product Description"]:
                        cell.font = Font(bold=False)
                cell.border = Border()
        wb.save(out_path)

        messagebox.showinfo(t["success"], t["success_msg"].format(path=out_path))

    except Exception as e:
        messagebox.showerror(t["error"], t["error_msg"].format(err=e))


# ------------------------------
# About dialog
# ------------------------------
def show_about():
    t = translations[current_lang.get()]
    messagebox.showinfo(
        t["about"],
        f"{APP_NAME}\n"
        f"Version {APP_VERSION} ({APP_DATE})\n"
        f"Created by {APP_AUTHOR}\n\n"
        f"Notes:\n{APP_NOTES}"
    )


# ------------------------------
# Load app info from JSON
# ------------------------------
base_dir = os.path.dirname(os.path.abspath(__file__))
info_path = resource_path("app_info.json")

try:
    with open(info_path, "r", encoding="utf-8") as f:
        app_info = json.load(f)
    APP_NAME = app_info.get("app_name", "My App")
    APP_VERSION = app_info.get("version", "0.0.0")
    APP_AUTHOR = app_info.get("author", "Unknown")
    APP_DATE = app_info.get("release_date", "")
    APP_NOTES = app_info.get("notes", "")
except Exception:
    APP_NAME = "My App"
    APP_VERSION = "0.0.0"
    APP_AUTHOR = "Unknown"
    APP_DATE = ""
    APP_NOTES = ""


# ------------------------------
# Tkinter GUI
# ------------------------------
root = tk.Tk()
root.title(f"{APP_NAME}")
root.geometry("700x500")
root.configure(bg="#f5f5f5")
root.minsize(700, 500)

# Now safe to create Tkinter variable
current_lang = tk.StringVar(value="en")  # default language

# Fonts
title_font = font.Font(family="Segoe UI", size=14, weight="bold")
normal_font = font.Font(family="Segoe UI", size=11)

# Language selector at top (centered group)
lang_frame = tk.Frame(root, bg="#f5f5f5")
lang_frame.pack(side="top", pady=10)

# Load flags (resize via subsample)
try:
    flag_en = PhotoImage(file=resource_path("assets/en.png")).subsample(20, 20)
    flag_hr = PhotoImage(file=resource_path("assets/hr.png")).subsample(20, 20)
except Exception:
    flag_en = None
    flag_hr = None

btn_en = tk.Button(
    lang_frame, image=flag_en, text="EN" if flag_en is None else "",
    compound="left",
    command=lambda: set_language("en"),
    borderwidth=0, cursor="hand2",
    bg="#f5f5f5", activebackground="#f5f5f5", padx=6, pady=4
)
btn_en.pack(side="left", padx=10)

btn_hr = tk.Button(
    lang_frame, image=flag_hr, text="HR" if flag_hr is None else "",
    compound="left",
    command=lambda: set_language("hr"),
    borderwidth=0, cursor="hand2",
    bg="#f5f5f5", activebackground="#f5f5f5", padx=6, pady=4
)
btn_hr.pack(side="left", padx=10)

# Center frame with card style
center_frame = tk.Frame(root, bg="white", padx=40, pady=30, relief="flat", bd=1)
center_frame.place(relx=0.5, rely=0.5, anchor="center")

title_label = tk.Label(center_frame, font=title_font, bg="white", wraplength=500, justify="center")
title_label.pack(pady=(0, 20))

upload_btn = tk.Button(
    center_frame, command=upload_structured, font=normal_font,
    bg="#4CAF50", fg="white", activebackground="#45a049",
    activeforeground="white", padx=15, pady=8,
    relief="flat", cursor="hand2"
)
upload_btn.pack(pady=8, fill="x")

file_label = tk.Label(
    center_frame,
    text="",  # set by set_language
    fg="gray", font=normal_font, bg="white"
)
file_label.pack(pady=5)

generate_btn = tk.Button(
    center_frame,
    command=process_file,
    font=normal_font,
    bg="#2196F3", fg="white",
    activebackground="#1976D2", activeforeground="white",
    padx=15, pady=8, relief="flat", cursor="hand2"
)
generate_btn.pack(pady=12, fill="x")

update_btn = tk.Button(
    center_frame, command=check_for_updates, font=normal_font,
    bg="#FF9800", fg="white", activebackground="#F57C00",
    activeforeground="white", padx=15, pady=8,
    relief="flat", cursor="hand2"
)
update_btn.pack(pady=12, fill="x")

# Footer
footer_frame = tk.Frame(root, bg="#f5f5f5")
footer_frame.pack(side="bottom", fill="x", pady=5)

footer_label = tk.Label(footer_frame, fg="gray", bg="#f5f5f5", font=("Segoe UI", 9))
footer_label.pack(side="left", padx=10)

about_btn = tk.Button(
    footer_frame,
    text="",  # set by set_language
    command=show_about,
    font=("Segoe UI", 9),
    bg="#e0e0e0",
    fg="black",
    relief="flat",
    padx=10,
    pady=2,
    cursor="hand2"
)
about_btn.pack(side="right", padx=10)

# Icon
system = platform.system()
try:
    if system == "Windows":
        root.iconbitmap(resource_path("assets/official-logo.ico"))
    else:
        root.iconphoto(False, PhotoImage(file=resource_path("assets/official-logo.png")))
except Exception:
    pass

# Initialize UI (Croatian was your last preference)
set_language("hr")

root.mainloop()
